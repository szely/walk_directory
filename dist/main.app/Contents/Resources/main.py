import os
from openpyxl import Workbook
from openpyxl.styles import PatternFill


def create_directory_structure_excel(directory, worksheet, row, col, indent=''):
    """
    Function to create the directory structure and list files recursively in an Excel file
    """
    if os.path.isdir(directory):
        # Highlight folder names in green
        folder_fill = PatternFill(start_color="00FF00", end_color="00FF00", fill_type="solid")
        worksheet.cell(row=row, column=col, value=os.path.basename(directory)).fill = folder_fill

        indent += '    '
        row += 1
        for item in sorted(os.listdir(directory)):
            if os.path.isdir(os.path.join(directory, item)):
                row = create_directory_structure_excel(os.path.join(directory, item), worksheet, row, col + 1, indent)
            else:
                worksheet.cell(row=row, column=col, value=os.path.basename(item))
                row += 1
    return row


# Specify the directory you want to start from
start_directory = '/Users/a1234/Yandex.Disk.localized/Документы/ОБУЧЕНИЕ/GB'

# Create a new workbook and select the active worksheet
wb = Workbook()
ws = wb.active

# Set the header
ws.cell(row=1, column=1, value='Directory Structure')

# Create the directory structure and list files in the Excel file
create_directory_structure_excel(start_directory, ws, row=2, col=1)

# Save the workbook
excel_file_path = 'directory_structure.xlsx'
wb.save(excel_file_path)
print(f"Directory structure saved to {excel_file_path}")
